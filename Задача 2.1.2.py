import csv, re, datetime, os

import matplotlib
import matplotlib.pyplot as plt
import numpy as np


from openpyxl.styles import Font, Side, Border
from openpyxl.workbook import Workbook
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00


class Report:
    def generate_excel(self, key_vac, statistics):
        wb = Workbook()
        thins = Side(border_style="thin", color="000000")
        sh1 = wb['Sheet']
        sh1.title = 'Статистика по годам'
        wb.create_sheet('Статистика по городам')
        col_names1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {key_vac}', 'Количество вакансий', f'Количество вакансий - {key_vac}']
        for i, column in enumerate(col_names1):
            sh1.cell(row=1, column=(i + 1), value=column).font = Font(bold=True)
        for year, value in statistics[0].items():
            sh1.append([year, value, statistics[1][year], statistics[2][year], statistics[3][year]])
        for column in sh1.columns:
            length = max(len(str(cell.value)) for cell in column)
            sh1.column_dimensions[column[0].column_letter].width = length + 1
            for cell in column:
                cell.border = Border(left=thins, top=thins, right=thins, bottom=thins)
        sh2 = wb['Статистика по городам']
        col_names2 = ['Город', 'Уровень зарплат', '  ', 'Город', 'Доля вакансий']
        for i, column in enumerate(col_names2):
            sh2.cell(row=1, column=(i + 1), value=column).font = Font(bold=True)
        activeStat = list(statistics[4].items())
        for i in range(10):
            activeStat[i] += tuple(statistics[5].items())[i]
        for city1, value1, city2, value2 in activeStat:
            sh2.append([city1, value1, '  ', city2, value2])
        for i in range(2, 12):
            sh2[f'E{i}'].number_format = FORMAT_PERCENTAGE_00
        for column in sh2.columns:
            length = max(len(str(cell.value)) for cell in column)
            sh2.column_dimensions[column[0].column_letter].width = length + 2
            for cell in column:
                if cell.value != '  ':
                    cell.border = Border(left=thins, top=thins, right=thins, bottom=thins)
        wb.save('report.xlsx')

    def generate_image(self, key_vac, statistics):
        matplotlib.rc('font', size=8)
        labels = statistics[0].keys()
        avg_salary = statistics[0].values()
        avg_salary_vac = statistics[1].values()
        count_vacancies = statistics[2].values()
        vacancies = statistics[3].values()
        cities = list(statistics[4].keys())
        salary_city = statistics[4].values()
        vac_city_names = list(statistics[5].values())
        vac_city_names = [1-sum(vac_city_names)] + vac_city_names
        vac_city =  ['Другие'] + list(statistics[5].keys())

        for i in range(len(cities)):
            cities[i] = cities[i].replace(' ', '\n')
            if '-' in cities[i]:
                cities[i] = '-\n'.join(cities[i].split('-'))

        x = np.arange(len(labels))
        width = 0.35

        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)
        ax1.bar(x - width / 2, avg_salary, width, label='средняя з/п')
        ax1.bar(x + width / 2, avg_salary_vac, width, label=f'з/п {key_vac}')
        ax1.set_title('Уровень зарплат по годам')
        ax1.set_xticks(x, labels, fontsize=8, rotation=90)
        ax1.legend(loc="upper left", fontsize=8)
        ax1.grid(axis='y')

        ax2.bar(x - width / 2, count_vacancies, width, label='Количество вакансий')
        ax2.bar(x + width / 2, vacancies, width, label=f'Количество вакансий {key_vac}')
        ax2.set_title('Количество вакансий по годам')
        ax2.set_xticks(x, labels, fontsize=8, rotation=90)
        ax2.legend(loc="upper left", fontsize=8)
        ax2.grid(axis='y')

        y_pos = np.arange(len(cities))
        ax3.barh(y_pos, salary_city, align='center')
        ax3.set_yticks(y_pos, labels=cities, fontsize=6)
        ax3.invert_yaxis()
        ax3.set_title('Уровень зарплат по городам')
        ax3.grid(axis='x')

        ax4.pie(vac_city_names, labels=vac_city, radius=1, textprops={"fontsize": 6})
        ax4.set_title('Доля вакансий по городам')
        fig.tight_layout()
        plt.savefig('graph.png')


class InputConect:
    def get_salary_level(self, list_vac, key, param_vac=''):
        result = {}
        for vacancy in list_vac:
            if vacancy.__getattribute__(key) not in result.keys():
                result[vacancy.__getattribute__(key)] = []
        list_vac = list((filter(lambda vac: param_vac in vac.name, list_vac))) if param_vac != '' else list_vac
        for vac in list_vac:
            result[vac.__getattribute__(key)].append(vac.salary.do_rub(float(vac.salary.salary_from) + float(vac.salary.salary_to)) / 2)
        for key in result.keys():
            result[key] = 0 if len(result[key]) == 0 else int(sum(result[key]) // len(result[key]))
        return result

    def get_count_vacancy(self, list_vac, key, dataSet, param_vac=''):
        result = {}
        for vacancy in list_vac:
            if vacancy.__getattribute__(key) not in result.keys():
                result[vacancy.__getattribute__(key)] = 0
        list_vac = list((filter(lambda vac: param_vac in vac.name, list_vac))) if param_vac != '' else list_vac
        for vac in list_vac:
            result[vac.__getattribute__(key)] += 1
        if key == 'area_name':
            for key in result.keys():
                result[key] = round(result[key] / len(dataSet.vacancies_objects), 4)
        return result

class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = [Vacancy(vac) for vac in self.csv_filer(*self.csv_reader(file_name))]

    def csv_filer(self, reader, list_naming):
        res = [self.formatting_str(x) for x in reader if len(x) == len(list_naming) and x.count('') == 0]
        dic_result = []
        for line in res:
            dic = {}
            for i in range(len(line)):
                dic[list_naming[i]] = line[i]
            dic_result.append(dic)
        return dic_result

    def csv_reader(self, file_name):
        file_csv = open(file_name, encoding='utf_8_sig')
        reader_csv = csv.reader(file_csv)
        list_data = []
        for str in reader_csv:
            list_data.append(str)
        columns = list_data[0]
        data = list_data[1:]
        return data, columns

    def formatting_str(self, element):
        for i in range(len(element)):
            element[i] = re.sub(r'<.*?>', '', element[i])
            if '\n' not in element[i]:
                element[i] = " ".join(element[i].split())
        return element


class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency

    def do_rub(self, salary):
        return salary * currency_to_rub[self.salary_currency]

class Vacancy:
    def __init__(self, data_vac):
        self.name = data_vac['name']
        self.salary = Salary(data_vac['salary_from'], data_vac['salary_to'], data_vac['salary_currency'])
        self.area_name = data_vac['area_name']
        self.published_at = data_vac['published_at']


currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

class OutputConnect:
    def check_published(self, date):
        return datetime.datetime.strptime(date, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y')

    def print_res(self, dataSet, outputer, vacancy_name):
        for vac in dataSet.vacancies_objects:
            vac.published_at = int(self.check_published(vac.published_at))
        res_year = dict(
            sorted(outputer.get_salary_level(dataSet.vacancies_objects, "published_at").items(), key=lambda x: x[0]))
        print(f'Динамика уровня зарплат по годам: {res_year}')
        res_vac = dict(
            sorted(outputer.get_count_vacancy(dataSet.vacancies_objects, "published_at", dataSet).items(), key=lambda x: x[0]))
        print(f'Динамика количества вакансий по годам: {res_vac}')
        res_year_vac = dict(sorted(outputer.get_salary_level(dataSet.vacancies_objects, "published_at", vacancy_name).items(),
                          key=lambda x: x[0]))
        print(f'Динамика уровня зарплат по годам для выбранной профессии: {res_year_vac}')
        res_vac_count = dict(sorted(outputer.get_count_vacancy(dataSet.vacancies_objects, "published_at", dataSet, vacancy_name).items(),
                          key=lambda x: x[0]))
        print(f'Динамика количества вакансий по годам для выбранной профессии: {res_vac_count}')
        dict_city_count = {}
        for vac in dataSet.vacancies_objects:
            if vac.area_name not in dict_city_count.keys():
                dict_city_count[vac.area_name] = 0
            dict_city_count[vac.area_name] += 1
        new_vacancies_objects = list(
            filter(lambda vac: int(len(dataSet.vacancies_objects) * 0.01) <= dict_city_count[vac.area_name],
                   dataSet.vacancies_objects))
        res_city = dict(sorted(outputer.get_salary_level(new_vacancies_objects, "area_name").items(), key=lambda x: x[1],
                          reverse=True)[:10])
        print(f'Уровень зарплат по городам (в порядке убывания): {res_city}')
        res_city_vac = dict(sorted(outputer.get_count_vacancy(new_vacancies_objects, "area_name", dataSet).items(), key=lambda x: x[1],
                          reverse=True)[:10])
        print(f'Доля вакансий по городам (в порядке убывания): {res_city_vac}')
        return [res_year, res_year_vac, res_vac, res_vac_count, res_city, res_city_vac]


file_name = input('Введите название файла: ')
vacancy_name = input('Введите название профессии: ')
if os.stat(file_name).st_size == 0:
    print("Пустой файл")
    exit()
outputer = InputConect()
dataSet = DataSet(file_name)
if len(dataSet.vacancies_objects) == 0:
    print('Нет данных')
    exit()
inputer = OutputConnect()
statistics = inputer.print_res(dataSet, outputer, vacancy_name)
rp = Report()
rp.generate_excel(vacancy_name, statistics)
rp.generate_image(vacancy_name, statistics)

